#Excelに設計したレンズの基本的な振る舞いを出力します


from re import A
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook


#等価屈折率を計算する関数
#媒質の屈折率n, 媒質の屈折率m, x方向の一次元占有率x, y方向の一次元屈折率y
def cal_tokakusseturitu(n, m, x, y):
    TE = x * n**2 + (1 - x) * m**2
    TM = y * n**(-2) + (1 - y) * m**(-2)

    ref = ((x * (1 / TM) + (1 - x) * m**2) / (y * (1 / TE) + (1 - y) * m**(-2)))**(1/4)
    return ref


#位相差を計算する関数
#中心からの距離r[m], 焦点距離f[m], 波長λ[m]
def cal_isousa(r, f, λ):
    φ = -360*(np.sqrt(r**2 + f**2) - f)/λ
    return φ


#中心との屈折率の差を計算する関数
#中心からの距離r[m], 焦点距離f[m], 媒質の高さH[m]
def cal_kusseturitusa(r, f, H):
    Δn = (np.sqrt(r**2 + f**2) - f) / H
    return Δn


#占有率と等価屈折率の関係を表す図を出力する関数
#大きい方の屈折率n, 小さい方の屈折率m
def print_relation_occandref(n, m):
    x = np.arange(0.0, 1.0, 0.01)
    ans = []
    for i in range(len(x)):
        ans.append(cal_tokakusseturitu(n, m, x[i], x[i]))

    plt.title('(Si:3.4, air:1)')
    plt.xlabel('occupancy')
    plt.ylabel('refractive index')
    plt.plot(x, ans, 'b-')
    plt.show()


#360°*nの位相差がつく距離はどこかを出力する関数
#基盤の辺の半分の長さradius[m], 焦点距離f[m], 波長λ[m]
def cal_r_isousa360(radius, f, λ, H):
    r_max = radius * np.sqrt(2)
    φ = -360
    r = 0
    while r_max > r:
        r = np.sqrt((f - φ*λ/360)**2 - f**2)
        print('位相差' + str(-φ) + '°つくのは中心から' + str(r*1000) + 'mmの位置')
        φ -= 360
        Δn = cal_kusseturitusa(r, f, H)
        print('またその位置での中心との屈折率の差Δnは' + str(Δn))


#位置rにおける中心との位相差、また占有率
def cal_isousa_occupancy(r, f, λ, H, silicon_ref):

    original_φ = -cal_isousa(r, f, λ)
    φ = -cal_isousa(r, f, λ)
    # print('中心との位相差は' + str(φ) + '°')
    while φ > 360:
        φ -= 360
    # print('360°調整した位相差は' + str(φ))
    Δn = (φ * λ / H) / 360
    # print('中心との屈折率差は' + str(Δn))
    Δn_360 = λ / H
    air_ref = 1.0
    # standard = (1.0 + silicon_ref) / 2   #基準
    standard = 1.8
    # standard = 1.5   #7.5用
    center_refrective = standard + Δn_360 / 2  #中心の等価屈折率
    r_refrective = center_refrective - Δn  #位置rでの等価屈折率
    # print('中心の等価屈折率は' + str(center_refrective))
    # print('位置rの等価屈折率は' + str(r_refrective))
    x = np.arange(0.0, 1.0, 0.001)
    occupancy = 0
    for i in range(len(x)):
        if cal_tokakusseturitu(silicon_ref, air_ref, x[i], x[i]) > r_refrective:
            occupancy = x[i]
            break
    # print('位置rの占有率は' + str(occupancy))

    return original_φ, φ, Δn, center_refrective, r_refrective, occupancy 

def create_excel(f, λ, H, silicon_ref, max_r):
    # wb = Workbook()
    # wb.save('myworkbook.xlsx')
    from openpyxl import load_workbook
    wb = load_workbook('myworkbook.xlsx')
    wb.create_sheet(index=0, title='exp6')
    ws = wb['exp6']
    # ws = wb.active
    ws.cell(row=2, column=3).value = '中心からの距離[mm]'
    ws.cell(row=2, column=4).value = '中心との位相差'
    ws.cell(row=2, column=5).value = '360°調整した位相差'
    ws.cell(row=2, column=6).value = '中心との屈折率差'
    ws.cell(row=2, column=7).value = '中心の等価屈折率'
    ws.cell(row=2, column=8).value = '位置rの等価屈折率'
    ws.cell(row=2, column=9).value = '位置rの占有率'
    y = np.arange(0.0, max_r, 0.2*10**(-3))
    for i in range(3, len(y)+3):
        original_φ, φ, Δn, center_refrective, r_refrective, occupancy = cal_isousa_occupancy(y[i-3], f, λ, H, silicon_ref)
        ws.cell(row=i, column=3).value = y[i-3]*10**3
        ws.cell(row=i, column=4).value = original_φ
        ws.cell(row=i, column=5).value = φ
        ws.cell(row=i, column=6).value = Δn
        ws.cell(row=i, column=7).value = center_refrective
        ws.cell(row=i, column=8).value = r_refrective
        ws.cell(row=i, column=9).value = occupancy
    wb.save('myworkbook.xlsx')
    wb.close()
        


create_excel(100*10**(-3), 750*10**(-6), 500*10**(-6), 3.4, 25.4*10**(-3))
# cal_isousa_occupancy(0.1*10**(-3), 100*10**(-3), 750*10**(-6), 400*10**(-6), 3.4)

    

